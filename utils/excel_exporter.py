import os
import openpyxl
from io import BytesIO


def generate_students_query_excel(students_data, template_path):
    """
    根据传入的学生数据，基于模板生成 Excel (保留 Sheet1，数据写入 Sheet2)
    """
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"未找到模板文件: {template_path}")

    # 读取现有模板（保留样式和结构）
    wb = openpyxl.load_workbook(template_path)

    # 强制操作第二个工作表 (Sheet2)
    sheet = wb.worksheets[1]

    # 按照模板 Sheet2 的表头顺序，将字典字段映射到对应列
    # [内部编号, 姓名*, 省学籍辅号, 国家身份证号, 性别, 校区*, 入学年份*, 当前班级*, 选科*, 外语种类*, 类别*, 专业, 在校情况*, 住宿情况*, 曾用名, 公寓, 宿舍及床位, 特殊情况备注]
    for row_idx, student in enumerate(students_data, start=2):  # 从第2行开始写
        sheet.cell(row=row_idx, column=1, value=student.get('custom_id', ''))
        sheet.cell(row=row_idx, column=2, value=student.get('name', ''))
        sheet.cell(row=row_idx, column=3, value=student.get('school_id', ''))
        sheet.cell(row=row_idx, column=4, value=student.get('national_id', ''))
        sheet.cell(row=row_idx, column=5, value=student.get('sex', ''))
        sheet.cell(row=row_idx, column=6, value=student.get('campus', ''))
        sheet.cell(row=row_idx, column=7, value=student.get('enter_year', ''))
        sheet.cell(row=row_idx, column=8, value=student.get('current_class', ''))
        sheet.cell(row=row_idx, column=9, value=student.get('subject', ''))
        sheet.cell(row=row_idx, column=10, value=student.get('language_type', ''))
        sheet.cell(row=row_idx, column=11, value=student.get('category', ''))
        sheet.cell(row=row_idx, column=12, value=student.get('major', ''))
        sheet.cell(row=row_idx, column=13, value=student.get('at_school', ''))
        sheet.cell(row=row_idx, column=14, value=student.get('boarding_status', ''))
        sheet.cell(row=row_idx, column=15, value=student.get('former_name', ''))
        sheet.cell(row=row_idx, column=16, value=student.get('apartment', ''))
        sheet.cell(row=row_idx, column=17, value=student.get('dormitory', ''))
        sheet.cell(row=row_idx, column=18, value=student.get('remarks', ''))

    # 将生成的 Excel 保存到内存的字节流中，避免在服务器上产生垃圾文件
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output